#Program reduces prices of items on a spreadsheet by 10%
import openpyxl as xl #xl is just an alias (not required) for easy typing
wb = xl.load_workbook('./venv/transactions.xlsx') #accesses the excel workbook
sheet = wb['Sheet1']                            #accesses sheet 1 of the workbook
#cell = sheet['a1']      #accesses our first cell a1. cell = sheet.cell(1,1) does exactly the same thing.
#print(cell.value)       #prints cell value "transaction_id"
#print(sheet.max_row)    #Tells us the max number of object rows on this spreadsheet. Results: 4

for row in range(2, sheet.max_row + 1):  # We exclude titles. Range typically excludes the last number. Hence we add 1
    cell = sheet.cell(row, 3)   #accesses column 3 element for each row.
    corrected_price = cell.value * 0.9   # gets the adjusted price for each col 3 element
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price

wb.save('./venv/transactions_corrected.xlsx')  #saves results in a corrected file.

# See the clean program with charts and more in pgm_excel2.py

