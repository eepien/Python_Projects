import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename) #./venv/transactions.xlsx
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col = 4, max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')   # We want to add the chart to E2

    wb.save(filename)  #saves results in the same file.
    # we could add more complexities like add label, legend, chart type,...etc
    # For more, read the documentation on open file excel.

#Now we are calling the function below
process_workbook('./venv/transactions.xlsx')